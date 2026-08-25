"""
Daily Habit & Task Reminder Telegram Bot
==========================================
Features:
  - Daily tasks with tappable check-in and streak tracking
  - Multiple reminder times per day (/addtime, /removetime, /mytimes)
  - Snooze button (delay a check-in by 1 hour)
  - Motivational quotes with each check-in
  - Streak milestone badges (7 / 30 / 100 days)
  - Excel export of your own data any time (/export)
  - Permanent storage on Render's free tier via Turso (optional, free)
  - Bilingual: Khmer (ខ្មែរ) and English, switch anytime with /language

Run:
    pip install -r requirements.txt
    python bot.py

Environment variable required:
    BOT_TOKEN   - your Telegram bot token from @BotFather

Optional (for permanent storage across Render redeploys — see README):
    TURSO_DATABASE_URL, TURSO_AUTH_TOKEN

Timezone:
    Defaults to Asia/Phnom_Penh (UTC+7). Change TIMEZONE_OFFSET_HOURS below if needed.
"""

import os
import io
import json
import random
import secrets
import sqlite3
import logging
import asyncio
import threading
from functools import wraps
from datetime import datetime, time, timedelta, timezone

from flask import Flask, request, redirect, url_for, session, render_template_string, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile, BotCommand
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

try:
    import libsql_experimental as libsql
except ImportError:
    libsql = None

# ------------------------------------------------------------------
# CONFIG
# ------------------------------------------------------------------
BOT_TOKEN = os.environ.get("BOT_TOKEN", "PUT_YOUR_TOKEN_HERE")
ADMIN_CHAT_ID = int(os.environ.get("ADMIN_CHAT_ID", "0"))  # your own Telegram chat_id (for /maintenance)
DASHBOARD_PASSWORD = os.environ.get("DASHBOARD_PASSWORD", "changeme")  # web dashboard login
FLASK_SECRET_KEY = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))
TIMEZONE_OFFSET_HOURS = 7  # Asia/Phnom_Penh (UTC+7), no DST
TZ = timezone(timedelta(hours=TIMEZONE_OFFSET_HOURS))
DB_PATH = os.path.join(os.path.dirname(__file__), "habit_bot.db")
DEFAULT_REMINDER_HOUR = 20
DEFAULT_REMINDER_MINUTE = 0
DEFAULT_LANGUAGE = "km"  # "km" = Khmer, "en" = English
BADGE_MILESTONES = (7, 30, 100)

# --- Permanent storage (free, via Turso) ---------------------------------
# Render's free tier wipes the local disk on every redeploy/restart. If
# TURSO_DATABASE_URL + TURSO_AUTH_TOKEN are set, habit_bot.db becomes an
# "embedded replica": a normal local SQLite file that reads/writes at local
# speed but is transparently synced to a free Turso cloud database, so a
# redeploy never loses data. Without them, it just falls back to plain
# local SQLite (same as before) and data resets on every redeploy.
TURSO_DATABASE_URL = os.environ.get("TURSO_DATABASE_URL", "").strip()
TURSO_AUTH_TOKEN = os.environ.get("TURSO_AUTH_TOKEN", "").strip()
USE_TURSO = bool(TURSO_DATABASE_URL and TURSO_AUTH_TOKEN and libsql is not None)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

if TURSO_DATABASE_URL and TURSO_AUTH_TOKEN and libsql is None:
    logger.warning("TURSO_DATABASE_URL/TURSO_AUTH_TOKEN are set but the 'libsql-experimental' "
                    "package isn't installed — falling back to local-only SQLite.")

# In-memory: pending (not-yet-confirmed) check-in state per chat_id
# { chat_id: { task_id: bool } }
pending_checkins = {}

# In-memory: which chat_ids are currently expected to type a plain-text
# reply for a menu action (e.g. the task name after tapping "Add Task"),
# and which action that reply is for. { chat_id: "addtask" | "addtime" }
awaiting_input = {}


# ------------------------------------------------------------------
# QUOTES
# ------------------------------------------------------------------
QUOTES = {
    "km": [
        "រាល់ថ្ងៃដែលអ្នកខិតខំ គឺជាជំហានមួយទៅមុខ។",
        "ភាពជោគជ័យមិនមែនចេញពីថ្ងៃណាមួយទេ វាចេញពីភាពស្ថិតស្ថេរ។",
        "កុំបោះបង់ streak ដែលអ្នកបានកសាងមក!",
        "ជំហានតូចៗប្រចាំថ្ងៃ នាំទៅរកលទ្ធផលធំ។",
        "ថ្ងៃនេះជាឱកាសមួយទៀតដើម្បីកាន់តែប្រសើរ។",
    ],
    "en": [
        "Every day you show up is a step forward.",
        "Success isn't from one day — it's from consistency.",
        "Don't break the streak you've built!",
        "Small daily steps lead to big results.",
        "Today is another chance to get a little better.",
    ],
}


def random_quote(lang):
    return random.choice(QUOTES.get(lang, QUOTES[DEFAULT_LANGUAGE]))


# ------------------------------------------------------------------
# TRANSLATIONS
# ------------------------------------------------------------------
TEXT = {
    "welcome_short": {
        "km": (
            "👋 សូមស្វាគមន៍មកកាន់ Daily Habit Bot!\n"
            "ម៉ោងជូនដំណឹង default គឺ {hour:02d}:{minute:02d} (ម៉ោងកម្ពុជា)។\n\n"
            "ប្រើប៊ូតុងខាងក្រោមដើម្បីចាប់ផ្តើម — មិនចាំបាច់វាយ command ទេ៖"
        ),
        "en": (
            "👋 Welcome to your Daily Habit Bot!\n"
            "Default reminder time is {hour:02d}:{minute:02d} (Phnom Penh time).\n\n"
            "Use the buttons below to get started — no need to type commands:"
        ),
    },
    "welcome": {
        "km": (
            "👋 សូមស្វាគមន៍មកកាន់ Daily Habit Bot!\n\n"
            "Commands:\n"
            "/addtask <ឈ្មោះ> - បន្ថែម task ថ្មី\n"
            "/removetask <ឈ្មោះ> - លុប task\n"
            "/mytasks - មើលបញ្ជី task\n"
            "/addtime HH:MM - បន្ថែមម៉ោងជូនដំណឹង (អាចមានច្រើន)\n"
            "/removetime HH:MM - លុបម៉ោងជូនដំណឹង\n"
            "/mytimes - មើលម៉ោងជូនដំណឹងទាំងអស់\n"
            "/checkin - សាកល្បង check-in ភ្លាមៗ\n"
            "/stats - របាយការណ៍ 7 ថ្ងៃចុងក្រោយ\n"
            "/export - ទាញយកទិន្នន័យខ្លួនឯង (Excel)\n"
            "/language - ប្តូរភាសា\n\n"
            "ម៉ោងជូនដំណឹង default គឺ {hour:02d}:{minute:02d} (ម៉ោងកម្ពុជា)។ ប្រើ /addtime ដើម្បីបន្ថែម។"
        ),
        "en": (
            "👋 Welcome to your Daily Habit Bot!\n\n"
            "Commands:\n"
            "/addtask <name> - add a task\n"
            "/removetask <name> - remove a task\n"
            "/mytasks - list your tasks\n"
            "/addtime HH:MM - add a reminder time (you can have several)\n"
            "/removetime HH:MM - remove a reminder time\n"
            "/mytimes - list all your reminder times\n"
            "/checkin - trigger a check-in now\n"
            "/stats - see your 7-day report\n"
            "/export - download your own data (Excel)\n"
            "/language - switch language\n\n"
            "Default reminder time is {hour:02d}:{minute:02d} (Phnom Penh time). Use /addtime to add more."
        ),
    },
    "addtask_usage": {"km": "របៀបប្រើ: /addtask <ឈ្មោះ task>", "en": "Usage: /addtask <task name>"},
    "addtask_duplicate": {"km": "⚠️ អ្នកមាន task ឈ្មោះ '{name}' រួចហើយ។", "en": "⚠️ You already have a task called '{name}'."},
    "addtask_success": {"km": "✅ បានបន្ថែម task: {name}", "en": "✅ Added daily task: {name}"},
    "removetask_usage": {"km": "របៀបប្រើ: /removetask <ឈ្មោះ task>", "en": "Usage: /removetask <task name>"},
    "removetask_notfound": {"km": "⚠️ រកមិនឃើញ task '{name}' ទេ។", "en": "⚠️ No task found named '{name}'."},
    "removetask_success": {"km": "🗑️ បានលុប task: {name}", "en": "🗑️ Removed task: {name}"},
    "mytasks_empty": {"km": "អ្នកមិនទាន់មាន task ទេ។ ប្រើ /addtask <ឈ្មោះ>", "en": "You have no tasks yet. Add one with /addtask <name>"},
    "mytasks_header": {"km": "📋 Task ប្រចាំថ្ងៃរបស់អ្នក:\n", "en": "📋 Your daily tasks:\n"},
    "mytasks_line": {"km": "• {name}  (streak: {streak} 🔥, ល្អបំផុត: {best})", "en": "• {name}  (streak: {streak} 🔥, best: {best})"},
    "addtime_usage": {"km": "របៀបប្រើ: /addtime HH:MM (ឧ. /addtime 07:30)", "en": "Usage: /addtime HH:MM (e.g. /addtime 07:30)"},
    "addtime_invalid": {"km": "⚠️ ម៉ោងមិនត្រឹមត្រូវ។ ប្រើទម្រង់ 24h ដូចជា 21:30", "en": "⚠️ Invalid time. Use 24h format like 21:30."},
    "addtime_exists": {"km": "⚠️ អ្នកមានម៉ោង {hour:02d}:{minute:02d} រួចហើយ។", "en": "⚠️ You already have {hour:02d}:{minute:02d} set."},
    "addtime_success": {"km": "⏰ បានបន្ថែមម៉ោងជូនដំណឹង: {hour:02d}:{minute:02d}", "en": "⏰ Added reminder time: {hour:02d}:{minute:02d}"},
    "removetime_usage": {"km": "របៀបប្រើ: /removetime HH:MM", "en": "Usage: /removetime HH:MM"},
    "removetime_notfound": {"km": "⚠️ រកមិនឃើញម៉ោង {hour:02d}:{minute:02d} ទេ។", "en": "⚠️ No reminder found at {hour:02d}:{minute:02d}."},
    "removetime_success": {"km": "🗑️ បានលុបម៉ោង: {hour:02d}:{minute:02d}", "en": "🗑️ Removed reminder time: {hour:02d}:{minute:02d}"},
    "mytimes_empty": {"km": "អ្នកមិនទាន់មានម៉ោងជូនដំណឹងទេ។ ប្រើ /addtime HH:MM", "en": "You have no reminder times yet. Add one with /addtime HH:MM"},
    "mytimes_header": {"km": "⏰ ម៉ោងជូនដំណឹងរបស់អ្នក:\n", "en": "⏰ Your reminder times:\n"},
    "stats_empty": {"km": "អ្នកមិនទាន់មាន task ទេ។ ប្រើ /addtask <ឈ្មោះ>", "en": "You have no tasks yet. Add one with /addtask <name>"},
    "stats_header": {"km": "📊 7 ថ្ងៃចុងក្រោយ:\n", "en": "📊 Last 7 days:\n"},
    "stats_line": {"km": "• {name}: បានធ្វើ {done} / កត់ត្រា {total} ({pct})", "en": "• {name}: {done} done / {total} logged ({pct})"},
    "checkin_prompt": {
        "km": "🌙 Check-in ប្រចាំថ្ងៃ — {date}\nចុចរាល់ task ដែលបានធ្វើថ្ងៃនេះ រួចចុច Confirm:\n\n💬 {quote}",
        "en": "🌙 Daily check-in — {date}\nTap each task you completed today, then Confirm:\n\n💬 {quote}",
    },
    "checkin_confirm_button": {"km": "✔️ បញ្ជាក់ check-in", "en": "✔️ Confirm check-in"},
    "checkin_snooze_button": {"km": "😴 រំលឹកម្តងទៀតក្នុង 1 ម៉ោង", "en": "😴 Snooze 1 hour"},
    "checkin_snoozed": {"km": "😴 បានពន្យារពេល — នឹងរំលឹកម្តងទៀតក្នុង 1 ម៉ោង។", "en": "😴 Snoozed — I'll remind you again in 1 hour."},
    "checkin_done": {"km": "🌙 Check-in ចប់ — {date}\n\n", "en": "🌙 Check-in complete — {date}\n\n"},
    "checkin_task_done": {"km": "✅ {name} — streak {streak} 🔥", "en": "✅ {name} — streak {streak} 🔥"},
    "checkin_task_reset": {"km": "❌ {name} — streak ត្រូវបានកំណត់ឡើងវិញ", "en": "❌ {name} — streak reset"},
    "badge_earned": {
        "km": "🏆 អបអរសាទរ! អ្នកទទួល badge '{badge}' សម្រាប់ task '{name}' (streak {streak} ថ្ងៃ)!",
        "en": "🏆 Congrats! You earned the '{badge}' badge for '{name}' ({streak}-day streak)!",
    },
    "language_prompt": {"km": "🌐 សូមជ្រើសរើសភាសា:", "en": "🌐 Choose your language:"},
    "language_set": {"km": "✅ ភាសាត្រូវបានប្តូរទៅជា ខ្មែរ", "en": "✅ Language switched to English"},
    "export_caption": {"km": "📊 ទិន្នន័យរបស់អ្នក (Excel)", "en": "📊 Your data export (Excel)"},
    "maintenance_active": {
        "km": "🛠 Bot កំពុងស្ថិតក្នុងការថែទាំបណ្តោះអាសន្ន សូមរង់ចាំបន្តិច ហើយសាកល្បងម្តងទៀតពេលក្រោយ។",
        "en": "🛠 The bot is under maintenance right now. Please try again shortly.",
    },
    "maintenance_usage": {"km": "របៀបប្រើ: /maintenance on | off | status", "en": "Usage: /maintenance on | off | status"},
    "maintenance_no_permission": {"km": "⛔ Command នេះសម្រាប់តែ admin ប៉ុណ្ណោះ។", "en": "⛔ This command is admin-only."},
    "maintenance_now_on": {"km": "🛠 Maintenance mode: ON — reminders និង commands ត្រូវផ្អាកសម្រាប់ user ធម្មតា។", "en": "🛠 Maintenance mode: ON — reminders and commands are paused for regular users."},
    "maintenance_now_off": {"km": "✅ Maintenance mode: OFF — bot ដំណើរការធម្មតាវិញ។", "en": "✅ Maintenance mode: OFF — bot is back to normal."},
    "maintenance_status": {"km": "Maintenance mode ឥឡូវនេះ: {state}", "en": "Maintenance mode is currently: {state}"},
    "menu_title": {"km": "📋 ជ្រើសរើសសកម្មភាព៖", "en": "📋 Choose an action:"},
    "menu_btn_mytasks": {"km": "📋 Task របស់ខ្ញុំ", "en": "📋 My Tasks"},
    "menu_btn_mytimes": {"km": "⏰ ម៉ោងរបស់ខ្ញុំ", "en": "⏰ My Times"},
    "menu_btn_stats": {"km": "📊 ស្ថិតិ", "en": "📊 Stats"},
    "menu_btn_checkin": {"km": "✅ Check-in ឥឡូវនេះ", "en": "✅ Check-in Now"},
    "menu_btn_export": {"km": "📥 ទាញយកទិន្នន័យ", "en": "📥 Export Data"},
    "menu_btn_language": {"km": "🌐 ភាសា", "en": "🌐 Language"},
    "menu_btn_addtask": {"km": "➕ បន្ថែម Task", "en": "➕ Add Task"},
    "menu_btn_removetask": {"km": "➖ លុប Task", "en": "➖ Remove Task"},
    "menu_btn_addtime": {"km": "⏰➕ បន្ថែមម៉ោង", "en": "⏰➕ Add Time"},
    "menu_btn_removetime": {"km": "⏰➖ លុបម៉ោង", "en": "⏰➖ Remove Time"},
    "menu_btn_help": {"km": "❓ ជំនួយ", "en": "❓ Help"},
    "menu_btn_back": {"km": "🔙 ត្រឡប់ទៅម៉ឺនុយ", "en": "🔙 Back to Menu"},
    "menu_open_hint": {"km": "ប្រើ /menu ដើម្បីមើលម៉ឺនុយម្តងទៀត។", "en": "Use /menu any time to see this again."},
    "btn_cancel": {"km": "❌ បោះបង់", "en": "❌ Cancel"},
    "input_cancelled": {"km": "❎ បានបោះបង់។", "en": "❎ Cancelled."},
    "addtask_prompt": {"km": "✏️ សូមវាយឈ្មោះ task ដែលអ្នកចង់បន្ថែម ហើយផ្ញើមកខ្ញុំ៖", "en": "✏️ Type the name of the task you'd like to add and send it to me:"},
    "removetask_pick": {"km": "🗑️ ចុចលើ task ដែលអ្នកចង់លុប៖", "en": "🗑️ Tap a task to remove it:"},
    "addtime_prompt": {"km": "⏰ សូមផ្ញើម៉ោង (ទម្រង់ HH:MM ដូចជា 07:30) ដែលអ្នកចង់បន្ថែម៖", "en": "⏰ Send the time you'd like to add, in HH:MM format (e.g. 07:30):"},
    "removetime_pick": {"km": "🗑️ ចុចលើម៉ោងដែលអ្នកចង់លុប៖", "en": "🗑️ Tap a reminder time to remove it:"},
    "mytasks_quick_header": {"km": "\n\n👉 ចុចលើ task ណាមួយដើម្បីសម្គាល់ថាបានធ្វើ/មិនទាន់ធ្វើសម្រាប់ថ្ងៃនេះ៖", "en": "\n\n👉 Tap any task to mark it done / not done for today:"},
}

BADGE_NAMES = {
    7: {"km": "ភ្លើងសប្តាហ៍", "en": "Week Streak"},
    30: {"km": "ខ្លាំងពិសេស", "en": "Monthly Master"},
    100: {"km": "តារាហ្វូង", "en": "Century Club"},
}


def t(lang, key, **kwargs):
    lang = lang if lang in ("km", "en") else DEFAULT_LANGUAGE
    template = TEXT[key].get(lang, TEXT[key][DEFAULT_LANGUAGE])
    return template.format(**kwargs) if kwargs else template


# ------------------------------------------------------------------
# DATABASE
# ------------------------------------------------------------------
class Row(dict):
    """Dict-style row (row["col"]) that also supports dict(row), used for
    both backends so the rest of the code never needs to know which one
    it's talking to."""
    pass


def _wrap_rows(cursor, raw_rows):
    cols = [d[0] for d in cursor.description] if cursor.description else []
    return [Row(zip(cols, r)) for r in raw_rows]


class _LibsqlCursor:
    """Makes a libsql_experimental cursor look like a sqlite3.Row cursor."""
    def __init__(self, cursor):
        self._cursor = cursor

    def execute(self, sql, params=()):
        self._cursor.execute(sql, params)
        return self

    def fetchone(self):
        row = self._cursor.fetchone()
        if row is None:
            return None
        return _wrap_rows(self._cursor, [row])[0]

    def fetchall(self):
        return _wrap_rows(self._cursor, self._cursor.fetchall())

    @property
    def description(self):
        return self._cursor.description


class _LibsqlConn:
    """Wraps a libsql_experimental (Turso) connection so it behaves like a
    sqlite3.Connection with row_factory=sqlite3.Row — libsql returns plain
    tuples and has no row_factory support, so every `row["col"]` access
    elsewhere in this file would otherwise break."""
    def __init__(self, conn):
        self._conn = conn
        self._dirty = False  # True once a write happens on this connection

    def execute(self, sql, params=()):
        # Only SELECT/PRAGMA are read-only; everything else (INSERT/UPDATE/
        # DELETE/CREATE/ALTER) marks this connection dirty so close() knows
        # a sync is actually needed. This avoids doing a network round-trip
        # to Turso Cloud on every single read (which used to happen on
        # *every* incoming message, since maintenance_guard reads settings
        # and language on each command).
        if not sql.lstrip().upper().startswith(("SELECT", "PRAGMA")):
            self._dirty = True
        return _LibsqlCursor(self._conn.execute(sql, params))

    def cursor(self):
        return _LibsqlCursor(self._conn.cursor())

    def commit(self):
        self._conn.commit()

    def close(self):
        if self._dirty:
            try:
                self._conn.sync()  # push local writes up to Turso Cloud
            except Exception as exc:
                logger.warning("Turso sync failed (will retry next connection): %s", exc)
        self._conn.close()


def get_conn():
    if USE_TURSO:
        conn = libsql.connect(DB_PATH, sync_url=TURSO_DATABASE_URL, auth_token=TURSO_AUTH_TOKEN)
        return _LibsqlConn(conn)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    if USE_TURSO:
        # Pull the latest data down from Turso Cloud before we start reading
        # or creating tables locally, so a fresh Render container always
        # starts from the newest synced copy instead of an empty file.
        try:
            conn = libsql.connect(DB_PATH, sync_url=TURSO_DATABASE_URL, auth_token=TURSO_AUTH_TOKEN)
            conn.sync()
            conn.close()
            logger.info("Synced latest data down from Turso Cloud.")
        except Exception as exc:
            logger.warning("Initial Turso sync failed — continuing with local data: %s", exc)

    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            chat_id INTEGER PRIMARY KEY,
            reminder_hour INTEGER DEFAULT 20,
            reminder_minute INTEGER DEFAULT 0,
            language TEXT DEFAULT 'km'
        )
    """)
    existing_cols = [row["name"] for row in c.execute("PRAGMA table_info(users)").fetchall()]
    if "language" not in existing_cols:
        c.execute("ALTER TABLE users ADD COLUMN language TEXT DEFAULT 'km'")

    c.execute("""
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chat_id INTEGER,
            name TEXT,
            streak INTEGER DEFAULT 0,
            best_streak INTEGER DEFAULT 0,
            last_done_date TEXT
        )
    """)
    task_cols = [row["name"] for row in c.execute("PRAGMA table_info(tasks)").fetchall()]
    if "last_badge" not in task_cols:
        c.execute("ALTER TABLE tasks ADD COLUMN last_badge INTEGER DEFAULT 0")

    c.execute("""
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER,
            date TEXT,
            done INTEGER
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS reminder_times (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chat_id INTEGER,
            hour INTEGER,
            minute INTEGER
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    conn.commit()

    # Migrate old single reminder_hour/reminder_minute into reminder_times, once.
    users = c.execute("SELECT chat_id, reminder_hour, reminder_minute FROM users").fetchall()
    for u in users:
        existing = c.execute(
            "SELECT id FROM reminder_times WHERE chat_id=? AND hour=? AND minute=?",
            (u["chat_id"], u["reminder_hour"], u["reminder_minute"]),
        ).fetchone()
        any_time = c.execute("SELECT id FROM reminder_times WHERE chat_id=?", (u["chat_id"],)).fetchone()
        if not existing and not any_time:
            c.execute(
                "INSERT INTO reminder_times (chat_id, hour, minute) VALUES (?, ?, ?)",
                (u["chat_id"], u["reminder_hour"], u["reminder_minute"]),
            )
    conn.commit()
    conn.close()


def get_user_language(chat_id):
    conn = get_conn()
    row = conn.execute("SELECT language FROM users WHERE chat_id=?", (chat_id,)).fetchone()
    conn.close()
    return row["language"] if row and row["language"] else DEFAULT_LANGUAGE


def ensure_user(chat_id, language=None):
    conn = get_conn()
    conn.execute(
        "INSERT OR IGNORE INTO users (chat_id, reminder_hour, reminder_minute, language) VALUES (?, ?, ?, ?)",
        (chat_id, DEFAULT_REMINDER_HOUR, DEFAULT_REMINDER_MINUTE, language or DEFAULT_LANGUAGE),
    )
    conn.commit()
    conn.close()


def get_setting(key, default=None):
    conn = get_conn()
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row["value"] if row else default


def set_setting(key, value):
    conn = get_conn()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value),
    )
    conn.commit()
    conn.close()


def is_maintenance_on():
    return get_setting("maintenance_mode", "false") == "true"


def is_admin(chat_id):
    return ADMIN_CHAT_ID != 0 and chat_id == ADMIN_CHAT_ID


def maintenance_guard(func):
    """Blocks a handler for non-admin users while maintenance mode is on."""
    @wraps(func)
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        chat_id = update.effective_chat.id
        if is_maintenance_on() and not is_admin(chat_id):
            lang = get_user_language(chat_id)
            if update.callback_query:
                await update.callback_query.answer()
                await update.callback_query.message.reply_text(t(lang, "maintenance_active"))
            else:
                await update.message.reply_text(t(lang, "maintenance_active"))
            return
        return await func(update, context, *args, **kwargs)
    return wrapper


def today_str():
    return datetime.now(TZ).strftime("%Y-%m-%d")


def yesterday_str():
    return (datetime.now(TZ) - timedelta(days=1)).strftime("%Y-%m-%d")


# ------------------------------------------------------------------
# COMMANDS
# ------------------------------------------------------------------
@maintenance_guard
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    awaiting_input.pop(chat_id, None)
    ensure_user(chat_id)

    conn = get_conn()
    has_time = conn.execute("SELECT id FROM reminder_times WHERE chat_id=?", (chat_id,)).fetchone()
    if not has_time:
        conn.execute(
            "INSERT INTO reminder_times (chat_id, hour, minute) VALUES (?, ?, ?)",
            (chat_id, DEFAULT_REMINDER_HOUR, DEFAULT_REMINDER_MINUTE),
        )
        conn.commit()
    conn.close()

    schedule_all_times_for_user(context.application, chat_id)

    lang = get_user_language(chat_id)
    greeting = t(lang, "welcome_short", hour=DEFAULT_REMINDER_HOUR, minute=DEFAULT_REMINDER_MINUTE)
    await update.message.reply_text(
        f"{greeting}\n\n{t(lang, 'menu_title')}", reply_markup=build_menu_keyboard(lang)
    )


@maintenance_guard
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """The full typed-command reference, kept for anyone who prefers
    typing over tapping. /start and the menu's 'Help' button both use
    the short, buttons-first welcome instead."""
    chat_id = update.effective_chat.id
    awaiting_input.pop(chat_id, None)
    lang = get_user_language(chat_id)
    await update.message.reply_text(
        t(lang, "welcome", hour=DEFAULT_REMINDER_HOUR, minute=DEFAULT_REMINDER_MINUTE),
        reply_markup=back_to_menu_keyboard(lang),
    )


def build_language_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🇰🇭 ខ្មែរ", callback_data="setlang:km"),
            InlineKeyboardButton("🇬🇧 English", callback_data="setlang:en"),
        ]
    ])


def build_menu_keyboard(lang):
    """The in-chat button menu — tapping most of these runs the action
    immediately; the ones that need typed input (task/time name) show a
    usage hint instead, since Telegram buttons can't pre-fill text."""
    rows = [
        [
            InlineKeyboardButton(t(lang, "menu_btn_mytasks"), callback_data="menu:mytasks"),
            InlineKeyboardButton(t(lang, "menu_btn_mytimes"), callback_data="menu:mytimes"),
        ],
        [
            InlineKeyboardButton(t(lang, "menu_btn_stats"), callback_data="menu:stats"),
            InlineKeyboardButton(t(lang, "menu_btn_checkin"), callback_data="menu:checkin"),
        ],
        [
            InlineKeyboardButton(t(lang, "menu_btn_export"), callback_data="menu:export"),
            InlineKeyboardButton(t(lang, "menu_btn_language"), callback_data="menu:language"),
        ],
        [
            InlineKeyboardButton(t(lang, "menu_btn_addtask"), callback_data="menu:addtask"),
            InlineKeyboardButton(t(lang, "menu_btn_removetask"), callback_data="menu:removetask"),
        ],
        [
            InlineKeyboardButton(t(lang, "menu_btn_addtime"), callback_data="menu:addtime"),
            InlineKeyboardButton(t(lang, "menu_btn_removetime"), callback_data="menu:removetime"),
        ],
        [InlineKeyboardButton(t(lang, "menu_btn_help"), callback_data="menu:help")],
    ]
    return InlineKeyboardMarkup(rows)


def back_to_menu_keyboard(lang):
    """A single 'Back to Menu' button, attached to most action replies so
    the user never has to type /menu to keep navigating."""
    return InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, "menu_btn_back"), callback_data="menu:show")]])


def cancel_keyboard(lang):
    """Shown while the bot is waiting for a typed reply (add task / add
    time), so the user can back out without sending anything."""
    return InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, "btn_cancel"), callback_data="cancel_input")]])


@maintenance_guard
async def menu_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    awaiting_input.pop(chat_id, None)
    lang = get_user_language(chat_id)
    await update.message.reply_text(t(lang, "menu_title"), reply_markup=build_menu_keyboard(lang))


@maintenance_guard
async def language_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    await update.message.reply_text(t(lang, "language_prompt"), reply_markup=build_language_keyboard())


def _add_task_core(chat_id, lang, name):
    """Shared by /addtask (typed args) and the menu's conversational
    'Add Task' flow. Returns (reply_text, success)."""
    name = (name or "").strip()
    if not name:
        return t(lang, "addtask_usage"), False

    conn = get_conn()
    existing = conn.execute("SELECT id FROM tasks WHERE chat_id=? AND name=?", (chat_id, name)).fetchone()
    if existing:
        conn.close()
        return t(lang, "addtask_duplicate", name=name), False

    conn.execute("INSERT INTO tasks (chat_id, name) VALUES (?, ?)", (chat_id, name))
    conn.commit()
    conn.close()
    return t(lang, "addtask_success", name=name), True


def _remove_task_core(chat_id, lang, name):
    """Shared by /removetask (typed args) and the tap-to-remove menu flow.
    Returns (reply_text, success)."""
    name = (name or "").strip()
    conn = get_conn()
    task = conn.execute("SELECT id FROM tasks WHERE chat_id=? AND name=?", (chat_id, name)).fetchone()
    if not task:
        conn.close()
        return t(lang, "removetask_notfound", name=name), False

    conn.execute("DELETE FROM tasks WHERE id=?", (task["id"],))
    conn.execute("DELETE FROM logs WHERE task_id=?", (task["id"],))
    conn.commit()
    conn.close()
    return t(lang, "removetask_success", name=name), True


@maintenance_guard
async def add_task(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    if not context.args:
        await update.message.reply_text(t(lang, "addtask_usage"))
        return
    name = " ".join(context.args).strip()
    msg, ok = _add_task_core(chat_id, lang, name)
    await update.message.reply_text(msg, reply_markup=back_to_menu_keyboard(lang) if ok else None)


@maintenance_guard
async def remove_task(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    if not context.args:
        await update.message.reply_text(t(lang, "removetask_usage"))
        return
    name = " ".join(context.args).strip()
    msg, ok = _remove_task_core(chat_id, lang, name)
    await update.message.reply_text(msg, reply_markup=back_to_menu_keyboard(lang) if ok else None)


def _mytasks_text(chat_id, lang):
    conn = get_conn()
    tasks = conn.execute("SELECT name, streak, best_streak FROM tasks WHERE chat_id=?", (chat_id,)).fetchall()
    conn.close()

    if not tasks:
        return t(lang, "mytasks_empty")

    lines = [t(lang, "mytasks_header")]
    for row in tasks:
        lines.append(t(lang, "mytasks_line", name=row["name"], streak=row["streak"], best=row["best_streak"]))
    return "\n".join(lines)


def empty_tasks_keyboard(lang):
    """Shown wherever there are no tasks yet — offers a direct 'Add Task'
    tap instead of leaving the person to type /addtask themselves."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, "menu_btn_addtask"), callback_data="menu:addtask")],
        [InlineKeyboardButton(t(lang, "menu_btn_back"), callback_data="menu:show")],
    ])


def empty_times_keyboard(lang):
    """Shown wherever there are no reminder times yet — offers a direct
    'Add Time' tap instead of leaving the person to type /addtime."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, "menu_btn_addtime"), callback_data="menu:addtime")],
        [InlineKeyboardButton(t(lang, "menu_btn_back"), callback_data="menu:show")],
    ])


def build_quicktask_keyboard(chat_id, lang):
    """One tappable button per task, showing today's ✅/⬜ status. Tapping
    a task immediately marks it done/not-done for today — no separate
    Confirm step needed, unlike the full /checkin flow."""
    conn = get_conn()
    tasks = conn.execute("SELECT id, name FROM tasks WHERE chat_id=?", (chat_id,)).fetchall()
    date = today_str()
    rows = []
    for row in tasks:
        log = conn.execute("SELECT done FROM logs WHERE task_id=? AND date=?", (row["id"], date)).fetchone()
        checked = bool(log["done"]) if log else False
        label = f"{'✅' if checked else '⬜'} {row['name']}"
        rows.append([InlineKeyboardButton(label, callback_data=f"qtoggle:{row['id']}")])
    conn.close()
    rows.append([InlineKeyboardButton(t(lang, "menu_btn_back"), callback_data="menu:show")])
    return InlineKeyboardMarkup(rows)


@maintenance_guard
async def my_tasks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    text = _mytasks_text(chat_id, lang)
    conn = get_conn()
    has_tasks = conn.execute("SELECT id FROM tasks WHERE chat_id=?", (chat_id,)).fetchone()
    conn.close()
    if has_tasks:
        text += t(lang, "mytasks_quick_header")
        await update.message.reply_text(text, reply_markup=build_quicktask_keyboard(chat_id, lang))
    else:
        await update.message.reply_text(text, reply_markup=empty_tasks_keyboard(lang))


def _parse_hhmm(text):
    hour, minute = map(int, text.split(":"))
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        raise ValueError("out of range")
    return hour, minute


def _add_time_core(chat_id, lang, text_in):
    """Shared by /addtime (typed args) and the menu's conversational
    'Add Time' flow. Returns (reply_text, success)."""
    text_in = (text_in or "").strip()
    if ":" not in text_in:
        return t(lang, "addtime_usage"), False
    try:
        hour, minute = _parse_hhmm(text_in)
    except ValueError:
        return t(lang, "addtime_invalid"), False

    conn = get_conn()
    existing = conn.execute(
        "SELECT id FROM reminder_times WHERE chat_id=? AND hour=? AND minute=?", (chat_id, hour, minute)
    ).fetchone()
    if existing:
        conn.close()
        return t(lang, "addtime_exists", hour=hour, minute=minute), False

    conn.execute("INSERT INTO reminder_times (chat_id, hour, minute) VALUES (?, ?, ?)", (chat_id, hour, minute))
    conn.commit()
    conn.close()
    return t(lang, "addtime_success", hour=hour, minute=minute), True


def _remove_time_core(chat_id, lang, hour, minute):
    """Shared by /removetime (typed args) and the tap-to-remove menu flow.
    Returns (reply_text, success)."""
    conn = get_conn()
    row = conn.execute(
        "SELECT id FROM reminder_times WHERE chat_id=? AND hour=? AND minute=?", (chat_id, hour, minute)
    ).fetchone()
    if not row:
        conn.close()
        return t(lang, "removetime_notfound", hour=hour, minute=minute), False

    conn.execute("DELETE FROM reminder_times WHERE id=?", (row["id"],))
    conn.commit()
    conn.close()
    return t(lang, "removetime_success", hour=hour, minute=minute), True


@maintenance_guard
async def add_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    ensure_user(chat_id, lang)
    if not context.args:
        await update.message.reply_text(t(lang, "addtime_usage"))
        return
    msg, ok = _add_time_core(chat_id, lang, context.args[0])
    if ok:
        schedule_all_times_for_user(context.application, chat_id)
    await update.message.reply_text(msg, reply_markup=back_to_menu_keyboard(lang) if ok else None)


@maintenance_guard
async def remove_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    if not context.args or ":" not in context.args[0]:
        await update.message.reply_text(t(lang, "removetime_usage"))
        return
    try:
        hour, minute = _parse_hhmm(context.args[0])
    except ValueError:
        await update.message.reply_text(t(lang, "addtime_invalid"))
        return
    msg, ok = _remove_time_core(chat_id, lang, hour, minute)
    if ok:
        schedule_all_times_for_user(context.application, chat_id)
    await update.message.reply_text(msg, reply_markup=back_to_menu_keyboard(lang) if ok else None)


def _mytimes_text(chat_id, lang):
    conn = get_conn()
    rows = conn.execute(
        "SELECT hour, minute FROM reminder_times WHERE chat_id=? ORDER BY hour, minute", (chat_id,)
    ).fetchall()
    conn.close()

    if not rows:
        return t(lang, "mytimes_empty")

    lines = [t(lang, "mytimes_header")]
    for row in rows:
        lines.append(f"• {row['hour']:02d}:{row['minute']:02d}")
    return "\n".join(lines)


@maintenance_guard
async def my_times(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    conn = get_conn()
    has_times = conn.execute("SELECT id FROM reminder_times WHERE chat_id=?", (chat_id,)).fetchone()
    conn.close()
    kb = back_to_menu_keyboard(lang) if has_times else empty_times_keyboard(lang)
    await update.message.reply_text(_mytimes_text(chat_id, lang), reply_markup=kb)


def _stats_text(chat_id, lang):
    conn = get_conn()
    tasks = conn.execute("SELECT id, name FROM tasks WHERE chat_id=?", (chat_id,)).fetchall()

    if not tasks:
        conn.close()
        return t(lang, "stats_empty")

    cutoff = (datetime.now(TZ) - timedelta(days=7)).strftime("%Y-%m-%d")
    lines = [t(lang, "stats_header")]
    for row in tasks:
        logs = conn.execute("SELECT done FROM logs WHERE task_id=? AND date>=?", (row["id"], cutoff)).fetchall()
        done_count = sum(1 for l in logs if l["done"])
        total = len(logs) if logs else 0
        pct = f"{(done_count/total*100):.0f}%" if total else "n/a"
        lines.append(t(lang, "stats_line", name=row["name"], done=done_count, total=total, pct=pct))
    conn.close()
    return "\n".join(lines)


@maintenance_guard
async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    await update.message.reply_text(_stats_text(chat_id, lang), reply_markup=back_to_menu_keyboard(lang))


def build_user_export_xlsx(chat_id):
    """Build an in-memory .xlsx workbook with the user's tasks + full log history."""
    conn = get_conn()
    tasks = conn.execute("SELECT * FROM tasks WHERE chat_id=?", (chat_id,)).fetchall()

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row_idx=1):
        for cell in ws[row_idx]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def autosize(ws):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)

    # --- Sheet 1: Tasks summary ---
    ws_tasks = wb.active
    ws_tasks.title = "Tasks"
    ws_tasks.append(["Task Name", "Current Streak", "Best Streak", "Total Check-ins", "Completed"])
    for task_row in tasks:
        logs = conn.execute("SELECT done FROM logs WHERE task_id=?", (task_row["id"],)).fetchall()
        total = len(logs)
        completed = sum(1 for l in logs if l["done"])
        ws_tasks.append([task_row["name"], task_row["streak"], task_row["best_streak"], total, completed])
    style_header(ws_tasks)
    ws_tasks.freeze_panes = "A2"
    autosize(ws_tasks)

    # --- Sheet 2: Full daily log history ---
    ws_logs = wb.create_sheet("Logs")
    ws_logs.append(["Task Name", "Date", "Done"])
    for task_row in tasks:
        logs = conn.execute("SELECT date, done FROM logs WHERE task_id=? ORDER BY date", (task_row["id"],)).fetchall()
        for l in logs:
            ws_logs.append([task_row["name"], l["date"], "Yes" if l["done"] else "No"])
    style_header(ws_logs)
    ws_logs.freeze_panes = "A2"
    autosize(ws_logs)

    conn.close()

    file_obj = io.BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    return file_obj


@maintenance_guard
async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    file_obj = build_user_export_xlsx(chat_id)
    file_obj.name = f"habit_data_{chat_id}_{today_str()}.xlsx"
    await update.message.reply_document(document=InputFile(file_obj), caption=t(lang, "export_caption"))


# ------------------------------------------------------------------
# CHECK-IN FLOW (inline buttons)
# ------------------------------------------------------------------
def build_checkin_keyboard(chat_id, lang):
    state = pending_checkins.get(chat_id, {})
    conn = get_conn()
    tasks = conn.execute("SELECT id, name FROM tasks WHERE chat_id=?", (chat_id,)).fetchall()
    conn.close()

    rows = []
    for row in tasks:
        checked = state.get(row["id"], False)
        label = f"{'✅' if checked else '⬜'} {row['name']}"
        rows.append([InlineKeyboardButton(label, callback_data=f"toggle:{row['id']}")])
    rows.append([InlineKeyboardButton(t(lang, "checkin_confirm_button"), callback_data="confirm")])
    rows.append([InlineKeyboardButton(t(lang, "checkin_snooze_button"), callback_data="snooze")])
    return InlineKeyboardMarkup(rows)


def _apply_checkin_for_task(conn, lang, task_id, done, date):
    """Writes today's log entry for one task and updates its streak/best
    streak/badges accordingly. Shared by the batch /checkin 'Confirm' flow
    and the single-tap quick-toggle on the My Tasks menu.
    Returns (summary_line, badge_lines) — summary_line is None if the task
    no longer exists (e.g. deleted mid-checkin)."""
    task_row = conn.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()
    if not task_row:
        return None, []

    existing_log = conn.execute("SELECT id FROM logs WHERE task_id=? AND date=?", (task_id, date)).fetchone()
    if existing_log:
        conn.execute("UPDATE logs SET done=? WHERE id=?", (1 if done else 0, existing_log["id"]))
    else:
        conn.execute("INSERT INTO logs (task_id, date, done) VALUES (?, ?, ?)", (task_id, date, 1 if done else 0))

    badge_lines = []
    if done:
        if task_row["last_done_date"] == date:
            # Already checked in today (e.g. a second reminder time firing
            # the same day, or re-tapping quick-toggle) — keep the streak
            # as-is instead of recomputing it.
            new_streak = task_row["streak"]
        elif task_row["last_done_date"] == yesterday_str():
            new_streak = task_row["streak"] + 1
        else:
            new_streak = 1
        best = max(new_streak, task_row["best_streak"])
        conn.execute(
            "UPDATE tasks SET streak=?, best_streak=?, last_done_date=? WHERE id=?",
            (new_streak, best, date, task_id),
        )
        summary = t(lang, "checkin_task_done", name=task_row["name"], streak=new_streak)

        last_badge = task_row["last_badge"] or 0
        for milestone in BADGE_MILESTONES:
            if new_streak >= milestone > last_badge:
                badge_name = BADGE_NAMES[milestone].get(lang, BADGE_NAMES[milestone]["en"])
                badge_lines.append(t(lang, "badge_earned", badge=badge_name, name=task_row["name"], streak=new_streak))
                conn.execute("UPDATE tasks SET last_badge=? WHERE id=?", (milestone, task_id))
    else:
        conn.execute("UPDATE tasks SET streak=0 WHERE id=?", (task_id,))
        summary = t(lang, "checkin_task_reset", name=task_row["name"])

    return summary, badge_lines


async def send_checkin(chat_id, context: ContextTypes.DEFAULT_TYPE):
    if is_maintenance_on() and not is_admin(chat_id):
        return  # don't ping users with reminders while the bot is under maintenance
    conn = get_conn()
    tasks = conn.execute("SELECT id FROM tasks WHERE chat_id=?", (chat_id,)).fetchall()
    conn.close()
    if not tasks:
        return

    lang = get_user_language(chat_id)
    pending_checkins[chat_id] = {row["id"]: False for row in tasks}
    await context.bot.send_message(
        chat_id=chat_id,
        text=t(lang, "checkin_prompt", date=today_str(), quote=random_quote(lang)),
        reply_markup=build_checkin_keyboard(chat_id, lang),
    )


@maintenance_guard
async def manual_checkin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await send_checkin(update.effective_chat.id, context)


async def maintenance_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    lang = get_user_language(chat_id)
    if not is_admin(chat_id):
        await update.message.reply_text(t(lang, "maintenance_no_permission"))
        return

    if not context.args:
        state = "ON" if is_maintenance_on() else "OFF"
        await update.message.reply_text(t(lang, "maintenance_status", state=state))
        return

    arg = context.args[0].lower()
    if arg == "on":
        set_setting("maintenance_mode", "true")
        await update.message.reply_text(t(lang, "maintenance_now_on"))
    elif arg == "off":
        set_setting("maintenance_mode", "false")
        await update.message.reply_text(t(lang, "maintenance_now_off"))
    elif arg == "status":
        state = "ON" if is_maintenance_on() else "OFF"
        await update.message.reply_text(t(lang, "maintenance_status", state=state))
    else:
        await update.message.reply_text(t(lang, "maintenance_usage"))


async def _snoozed_checkin_job(context: ContextTypes.DEFAULT_TYPE):
    chat_id = context.job.chat_id
    await send_checkin(chat_id, context)


@maintenance_guard
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    chat_id = query.message.chat_id
    await query.answer()
    lang = get_user_language(chat_id)

    if query.data.startswith("setlang:"):
        new_lang = query.data.split(":")[1]
        ensure_user(chat_id, new_lang)
        conn = get_conn()
        conn.execute("UPDATE users SET language=? WHERE chat_id=?", (new_lang, chat_id))
        conn.commit()
        conn.close()
        await query.edit_message_text(t(new_lang, "language_set"))
        await context.bot.send_message(chat_id=chat_id, text=t(new_lang, "menu_title"), reply_markup=build_menu_keyboard(new_lang))
        return

    if query.data == "cancel_input":
        awaiting_input.pop(chat_id, None)
        await query.edit_message_text(t(lang, "input_cancelled"))
        await context.bot.send_message(chat_id=chat_id, text=t(lang, "menu_title"), reply_markup=build_menu_keyboard(lang))
        return

    if query.data.startswith("rmtask:"):
        task_id = int(query.data.split(":")[1])
        conn = get_conn()
        task = conn.execute("SELECT name FROM tasks WHERE id=? AND chat_id=?", (task_id, chat_id)).fetchone()
        conn.close()
        if task:
            msg, _ok = _remove_task_core(chat_id, lang, task["name"])
        else:
            msg = t(lang, "removetask_notfound", name="")
        await query.edit_message_text(msg, reply_markup=back_to_menu_keyboard(lang))
        return

    if query.data.startswith("rmtime:"):
        _, hour_s, minute_s = query.data.split(":")
        hour, minute = int(hour_s), int(minute_s)
        msg, ok = _remove_time_core(chat_id, lang, hour, minute)
        if ok:
            schedule_all_times_for_user(context.application, chat_id)
        await query.edit_message_text(msg, reply_markup=back_to_menu_keyboard(lang))
        return

    if query.data.startswith("qtoggle:"):
        task_id = int(query.data.split(":")[1])
        conn = get_conn()
        task_owned = conn.execute("SELECT id FROM tasks WHERE id=? AND chat_id=?", (task_id, chat_id)).fetchone()
        if not task_owned:
            conn.close()
            return
        date = today_str()
        log = conn.execute("SELECT done FROM logs WHERE task_id=? AND date=?", (task_id, date)).fetchone()
        currently_done = bool(log["done"]) if log else False
        summary, badge_lines = _apply_checkin_for_task(conn, lang, task_id, not currently_done, date)
        conn.commit()
        conn.close()
        if summary is not None:
            try:
                await query.edit_message_reply_markup(reply_markup=build_quicktask_keyboard(chat_id, lang))
            except Exception:
                pass  # message unchanged (e.g. double-tap race) — safe to ignore
            for badge_msg in badge_lines:
                await context.bot.send_message(chat_id=chat_id, text=badge_msg)
        return

    if query.data.startswith("menu:"):
        action = query.data.split(":", 1)[1]
        if action == "show":
            awaiting_input.pop(chat_id, None)
            await context.bot.send_message(chat_id=chat_id, text=t(lang, "menu_title"), reply_markup=build_menu_keyboard(lang))
        elif action == "mytasks":
            conn = get_conn()
            has_tasks = conn.execute("SELECT id FROM tasks WHERE chat_id=?", (chat_id,)).fetchone()
            conn.close()
            if has_tasks:
                text = _mytasks_text(chat_id, lang) + t(lang, "mytasks_quick_header")
                await context.bot.send_message(chat_id=chat_id, text=text, reply_markup=build_quicktask_keyboard(chat_id, lang))
            else:
                await context.bot.send_message(chat_id=chat_id, text=t(lang, "mytasks_empty"), reply_markup=empty_tasks_keyboard(lang))
        elif action == "mytimes":
            conn = get_conn()
            has_times = conn.execute("SELECT id FROM reminder_times WHERE chat_id=?", (chat_id,)).fetchone()
            conn.close()
            kb = back_to_menu_keyboard(lang) if has_times else empty_times_keyboard(lang)
            await context.bot.send_message(chat_id=chat_id, text=_mytimes_text(chat_id, lang), reply_markup=kb)
        elif action == "stats":
            await context.bot.send_message(chat_id=chat_id, text=_stats_text(chat_id, lang), reply_markup=back_to_menu_keyboard(lang))
        elif action == "checkin":
            await send_checkin(chat_id, context)
        elif action == "export":
            file_obj = build_user_export_xlsx(chat_id)
            file_obj.name = f"habit_data_{chat_id}_{today_str()}.xlsx"
            await context.bot.send_document(chat_id=chat_id, document=InputFile(file_obj), caption=t(lang, "export_caption"))
        elif action == "language":
            await context.bot.send_message(chat_id=chat_id, text=t(lang, "language_prompt"), reply_markup=build_language_keyboard())
        elif action == "help":
            await context.bot.send_message(chat_id=chat_id, text=t(lang, "welcome", hour=DEFAULT_REMINDER_HOUR, minute=DEFAULT_REMINDER_MINUTE), reply_markup=back_to_menu_keyboard(lang))
        elif action == "addtask":
            awaiting_input[chat_id] = "addtask"
            await context.bot.send_message(chat_id=chat_id, text=t(lang, "addtask_prompt"), reply_markup=cancel_keyboard(lang))
        elif action == "addtime":
            awaiting_input[chat_id] = "addtime"
            await context.bot.send_message(chat_id=chat_id, text=t(lang, "addtime_prompt"), reply_markup=cancel_keyboard(lang))
        elif action == "removetask":
            conn = get_conn()
            tasks = conn.execute("SELECT id, name FROM tasks WHERE chat_id=?", (chat_id,)).fetchall()
            conn.close()
            if not tasks:
                await context.bot.send_message(chat_id=chat_id, text=t(lang, "mytasks_empty"), reply_markup=empty_tasks_keyboard(lang))
            else:
                rows = [[InlineKeyboardButton(f"🗑️ {row['name']}", callback_data=f"rmtask:{row['id']}")] for row in tasks]
                rows.append([InlineKeyboardButton(t(lang, "btn_cancel"), callback_data="menu:show")])
                await context.bot.send_message(chat_id=chat_id, text=t(lang, "removetask_pick"), reply_markup=InlineKeyboardMarkup(rows))
        elif action == "removetime":
            conn = get_conn()
            times = conn.execute("SELECT hour, minute FROM reminder_times WHERE chat_id=? ORDER BY hour, minute", (chat_id,)).fetchall()
            conn.close()
            if not times:
                await context.bot.send_message(chat_id=chat_id, text=t(lang, "mytimes_empty"), reply_markup=empty_times_keyboard(lang))
            else:
                rows = [[InlineKeyboardButton(f"🗑️ {row['hour']:02d}:{row['minute']:02d}", callback_data=f"rmtime:{row['hour']}:{row['minute']}")] for row in times]
                rows.append([InlineKeyboardButton(t(lang, "btn_cancel"), callback_data="menu:show")])
                await context.bot.send_message(chat_id=chat_id, text=t(lang, "removetime_pick"), reply_markup=InlineKeyboardMarkup(rows))
        return

    if query.data == "snooze":
        context.job_queue.run_once(_snoozed_checkin_job, when=3600, chat_id=chat_id, name=f"snooze_{chat_id}_{datetime.now(TZ).timestamp()}")
        pending_checkins.pop(chat_id, None)
        await query.edit_message_text(t(lang, "checkin_snoozed"))
        return

    if chat_id not in pending_checkins:
        pending_checkins[chat_id] = {}

    if query.data.startswith("toggle:"):
        task_id = int(query.data.split(":")[1])
        current = pending_checkins[chat_id].get(task_id, False)
        pending_checkins[chat_id][task_id] = not current
        await query.edit_message_reply_markup(reply_markup=build_checkin_keyboard(chat_id, lang))

    elif query.data == "confirm":
        state = pending_checkins.get(chat_id, {})
        conn = get_conn()
        summary_lines = []
        badge_lines = []
        date = today_str()
        for task_id, done in state.items():
            summary, badges = _apply_checkin_for_task(conn, lang, task_id, done, date)
            if summary is None:
                continue
            summary_lines.append(summary)
            badge_lines.extend(badges)

        conn.commit()
        conn.close()
        pending_checkins.pop(chat_id, None)

        await query.edit_message_text(t(lang, "checkin_done", date=date) + "\n".join(summary_lines), reply_markup=back_to_menu_keyboard(lang))
        for badge_msg in badge_lines:
            await context.bot.send_message(chat_id=chat_id, text=badge_msg)


@maintenance_guard
async def handle_text_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handles a plain-text reply while the menu is waiting on one (e.g.
    the task name after tapping 'Add Task'). Ignores any other stray text
    — this bot only reacts to commands/buttons otherwise."""
    chat_id = update.effective_chat.id
    if chat_id not in awaiting_input:
        return
    lang = get_user_language(chat_id)
    action = awaiting_input.pop(chat_id)
    text_in = (update.message.text or "").strip()

    if action == "addtask":
        msg, ok = _add_task_core(chat_id, lang, text_in)
    elif action == "addtime":
        msg, ok = _add_time_core(chat_id, lang, text_in)
        if ok:
            schedule_all_times_for_user(context.application, chat_id)
    else:
        return

    await update.message.reply_text(msg, reply_markup=back_to_menu_keyboard(lang) if ok else cancel_keyboard(lang))
    if not ok:
        # invalid input (e.g. bad time format) — keep waiting for another try
        awaiting_input[chat_id] = action


# ------------------------------------------------------------------
# SCHEDULING
# ------------------------------------------------------------------
def schedule_all_times_for_user(application: Application, chat_id: int):
    for job in application.job_queue.get_jobs_by_name(f"reminder_{chat_id}_all"):
        job.schedule_removal()
    # remove any previously-scheduled per-time jobs for this user
    for job in list(application.job_queue.jobs()):
        if job.name and job.name.startswith(f"reminder_{chat_id}_"):
            job.schedule_removal()

    conn = get_conn()
    times = conn.execute("SELECT hour, minute FROM reminder_times WHERE chat_id=?", (chat_id,)).fetchall()
    conn.close()

    for row in times:
        job_name = f"reminder_{chat_id}_{row['hour']:02d}{row['minute']:02d}"
        application.job_queue.run_daily(
            callback=lambda ctx: send_checkin(chat_id, ctx),
            time=time(hour=row["hour"], minute=row["minute"], tzinfo=TZ),
            name=job_name,
            chat_id=chat_id,
        )


async def schedule_all_users(application: Application):
    conn = get_conn()
    chat_ids = [row["chat_id"] for row in conn.execute("SELECT DISTINCT chat_id FROM reminder_times").fetchall()]
    conn.close()
    for chat_id in chat_ids:
        schedule_all_times_for_user(application, chat_id)


# ------------------------------------------------------------------
# WEB DASHBOARD (control panel + UptimeRobot keep-alive on Render free tier)
# ------------------------------------------------------------------
flask_app = Flask(__name__)
flask_app.secret_key = FLASK_SECRET_KEY

DASHBOARD_CSS = """
:root{
  --bg:#0a0c10; --bg-glow:#12162080;
  --card:#141821; --card-hi:#171c27; --border:#232838; --border-hi:#31384c;
  --text:#eef0f5; --muted:#8890a4; --muted-dim:#5c637a;
  --accent:#6d8bff; --accent-2:#8f6dff; --good:#3ddc97; --bad:#ff6b81; --warn:#ffbd5c;
  --radius:16px;
}
*{box-sizing:border-box;}
body{
  margin:0; background:
    radial-gradient(1100px 500px at 15% -10%, var(--bg-glow), transparent 60%),
    radial-gradient(900px 500px at 100% 0%, #6d8bff14, transparent 55%),
    var(--bg);
  color:var(--text); min-height:100vh;
  font-family:"Inter",-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
}
.wrap{max-width:1040px;margin:0 auto;padding:36px 20px 64px;}
.topbar{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;margin-bottom:26px;}
.brand{display:flex;align-items:center;gap:12px;}
.brand .logo{
  width:42px;height:42px;border-radius:12px;display:flex;align-items:center;justify-content:center;
  font-size:20px;background:linear-gradient(135deg,var(--accent),var(--accent-2));
  box-shadow:0 6px 18px -6px #6d8bff70;
}
h1{font-size:20px;margin:0 0 2px;font-weight:700;letter-spacing:-.01em;}
.sub{color:var(--muted);font-size:13px;margin:0;}
.logout-link{color:var(--muted);text-decoration:none;font-size:13px;padding:8px 14px;border:1px solid var(--border);border-radius:10px;transition:.15s;}
.logout-link:hover{color:var(--text);border-color:var(--border-hi);}
.card{
  background:linear-gradient(180deg,var(--card-hi),var(--card));
  border:1px solid var(--border);border-radius:var(--radius);padding:22px;margin-bottom:16px;
  box-shadow:0 1px 0 #ffffff05 inset;
}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:16px;}
.grid2{display:grid;grid-template-columns:1fr 1.25fr;gap:16px;margin-bottom:16px;}
@media (max-width:720px){.grid2{grid-template-columns:1fr;}}
.stat{
  background:linear-gradient(180deg,var(--card-hi),var(--card));
  border:1px solid var(--border);border-radius:var(--radius);padding:18px;
}
.stat .n{font-size:28px;font-weight:800;letter-spacing:-.02em;background:linear-gradient(135deg,#fff,#b9c2e0);-webkit-background-clip:text;background-clip:text;color:transparent;}
.stat .l{color:var(--muted);font-size:12px;margin-top:6px;font-weight:500;}
.badge{display:inline-flex;align-items:center;gap:7px;padding:6px 14px;border-radius:999px;font-size:13px;font-weight:600;border:1px solid transparent;}
.badge.on{background:#ff6b8117;color:var(--bad);border-color:#ff6b8130;}
.badge.off{background:#3ddc9717;color:var(--good);border-color:#3ddc9730;}
.badge.cloud{background:#6d8bff17;color:var(--accent);border-color:#6d8bff30;}
.badge.local{background:#ffbd5c17;color:var(--warn);border-color:#ffbd5c30;}
.badge-row{display:flex;gap:10px;flex-wrap:wrap;}
.row{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;}
button{
  background:linear-gradient(135deg,var(--accent),var(--accent-2));color:#fff;border:none;border-radius:11px;
  padding:11px 20px;font-size:14px;font-weight:600;cursor:pointer;transition:.15s;box-shadow:0 4px 14px -6px #6d8bff60;
}
button:hover{filter:brightness(1.08);transform:translateY(-1px);}
button.danger{background:linear-gradient(135deg,#ff6b81,#ff4d6d);box-shadow:0 4px 14px -6px #ff6b8160;}
button.ghost{background:transparent;border:1px solid var(--border-hi);color:var(--text);box-shadow:none;}
table{width:100%;border-collapse:collapse;font-size:13px;}
th{text-align:left;color:var(--muted-dim);font-weight:600;padding:10px 12px;border-bottom:1px solid var(--border);
   text-transform:uppercase;font-size:11px;letter-spacing:.04em;}
td{padding:10px 12px;border-bottom:1px solid var(--border);}
tr:last-child td{border-bottom:none;}
tr:hover td{background:#ffffff03;}
.empty{color:var(--muted);font-size:13px;padding:10px 0;}
h2{font-size:13px;margin:0 0 16px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;font-weight:700;}
input[type=password]{
  width:100%;padding:13px 14px;border-radius:11px;border:1px solid var(--border);
  background:#0d1017;color:var(--text);font-size:14px;margin-bottom:14px;outline:none;transition:.15s;
}
input[type=password]:focus{border-color:var(--accent);}
.login-box{max-width:360px;margin:100px auto 0;}
.login-box .card{text-align:center;}
.login-box .logo{width:52px;height:52px;font-size:24px;margin:0 auto 16px;}
.err{color:var(--bad);font-size:13px;margin-bottom:10px;background:#ff6b8112;padding:8px 12px;border-radius:8px;}
a{color:var(--accent);text-decoration:none;}
"""

LOGIN_HTML = """
<!doctype html><html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<title>Habit Bot — Login</title><style>{{ css }}</style></head>
<body><div class="wrap login-box"><div class="card">
<div class="logo" style="border-radius:14px;display:flex;align-items:center;justify-content:center;background:linear-gradient(135deg,var(--accent),var(--accent-2));box-shadow:0 6px 18px -6px #6d8bff70;">🔒</div>
<h1>Habit Bot Dashboard</h1><p class="sub" style="margin-bottom:22px;">Enter the admin password to continue.</p>
{% if error %}<p class="err">{{ error }}</p>{% endif %}
<form method="post"><input type="password" name="password" placeholder="Password" autofocus>
<button type="submit" style="width:100%">Log in</button></form>
</div></div></body></html>
"""

DASHBOARD_HTML = """
<!doctype html><html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<title>Habit Bot — Dashboard</title><style>{{ css }}</style></head>
<body><div class="wrap">

<div class="topbar">
  <div class="brand">
    <div class="logo">📊</div>
    <div><h1>Daily Habit Bot</h1><p class="sub">Control panel &amp; live stats</p></div>
  </div>
  <a class="logout-link" href="{{ url_for('logout') }}">Log out</a>
</div>

<div class="card row">
  <div class="badge-row">
    <span class="badge {{ 'on' if maintenance else 'off' }}">
      {{ '🛠 Maintenance ON' if maintenance else '✅ Running normally' }}
    </span>
    <span class="badge {{ 'cloud' if db_synced else 'local' }}">
      {{ '☁️ Turso Cloud — data is permanent' if db_synced else '💾 Local only — data resets on redeploy' }}
    </span>
  </div>
  <form method="post" action="{{ url_for('toggle_maintenance') }}">
    <input type="hidden" name="csrf_token" value="{{ csrf_token }}">
    <button class="{{ 'ghost' if maintenance else 'danger' }}" type="submit">
      {{ 'Turn maintenance OFF' if maintenance else 'Turn maintenance ON' }}
    </button>
  </form>
</div>

<div class="grid">
  <div class="stat"><div class="n">{{ stats.users }}</div><div class="l">Total users</div></div>
  <div class="stat"><div class="n">{{ stats.tasks }}</div><div class="l">Total tasks</div></div>
  <div class="stat"><div class="n">{{ stats.checkins_today }}</div><div class="l">Check-ins today</div></div>
  <div class="stat"><div class="n">{{ stats.checkins_week }}</div><div class="l">Check-ins (7d)</div></div>
</div>

<div class="card">
  <h2>Check-ins — last 14 days</h2>
  <canvas id="checkinChart" height="90"></canvas>
</div>

<div class="grid2">
  <div class="card">
    <h2>Language split</h2>
    {% if users %}
    <canvas id="langChart" height="180"></canvas>
    {% else %}<p class="empty">No users yet.</p>{% endif %}
  </div>

  <div class="card">
    <h2>Top streaks</h2>
    {% if top_streaks %}
    <table><tr><th>Task</th><th>User (chat_id)</th><th>Streak</th><th>Best</th></tr>
    {% for r in top_streaks %}
    <tr><td>{{ r.name }}</td><td>{{ r.chat_id }}</td><td>🔥 {{ r.streak }}</td><td>{{ r.best_streak }}</td></tr>
    {% endfor %}</table>
    {% else %}<p class="empty">No tasks yet.</p>{% endif %}
  </div>
</div>

<div class="card">
  <h2>Users</h2>
  {% if users %}
  <table><tr><th>chat_id</th><th>Language</th><th>Tasks</th><th>Reminder times</th></tr>
  {% for u in users %}
  <tr><td>{{ u.chat_id }}</td><td>{{ u.language }}</td><td>{{ u.task_count }}</td><td>{{ u.times }}</td></tr>
  {% endfor %}</table>
  {% else %}<p class="empty">No users yet.</p>{% endif %}
</div>

</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.4/chart.umd.min.js"></script>
<script>
  const gridColor = "#232838", textColor = "#8890a4";
  Chart.defaults.color = textColor;
  Chart.defaults.borderColor = gridColor;
  Chart.defaults.font.family = "Inter, sans-serif";

  const checkinData = {{ checkin_series_json | safe }};
  new Chart(document.getElementById('checkinChart'), {
    type: 'bar',
    data: {
      labels: checkinData.labels,
      datasets: [
        { label: 'Done', data: checkinData.done, backgroundColor: '#3ddc97', borderRadius: 4, stack: 's' },
        { label: 'Missed', data: checkinData.missed, backgroundColor: '#ff6b81', borderRadius: 4, stack: 's' }
      ]
    },
    options: {
      responsive: true,
      scales: { x: { stacked: true, grid: { display: false } }, y: { stacked: true, beginAtZero: true, ticks: { precision: 0 } } },
      plugins: { legend: { position: 'bottom' } }
    }
  });

  {% if users %}
  const langData = {{ lang_breakdown_json | safe }};
  new Chart(document.getElementById('langChart'), {
    type: 'doughnut',
    data: {
      labels: langData.labels,
      datasets: [{ data: langData.counts, backgroundColor: ['#6d8bff', '#ffbd5c', '#3ddc97'], borderWidth: 0 }]
    },
    options: { responsive: true, cutout: '65%', plugins: { legend: { position: 'bottom' } } }
  });
  {% endif %}
</script>
</body></html>
"""


def _require_login(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapper


def _dashboard_stats():
    conn = get_conn()
    users = conn.execute("SELECT COUNT(*) c FROM users").fetchone()["c"]
    tasks = conn.execute("SELECT COUNT(*) c FROM tasks").fetchone()["c"]
    checkins_today = conn.execute(
        "SELECT COUNT(*) c FROM logs WHERE date=? AND done=1", (today_str(),)
    ).fetchone()["c"]
    week_cutoff = (datetime.now(TZ) - timedelta(days=7)).strftime("%Y-%m-%d")
    checkins_week = conn.execute(
        "SELECT COUNT(*) c FROM logs WHERE date>=? AND done=1", (week_cutoff,)
    ).fetchone()["c"]
    conn.close()
    return {"users": users, "tasks": tasks, "checkins_today": checkins_today, "checkins_week": checkins_week}


def _top_streaks(limit=10):
    conn = get_conn()
    rows = conn.execute(
        "SELECT chat_id, name, streak, best_streak FROM tasks ORDER BY streak DESC, best_streak DESC LIMIT ?",
        (limit,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _user_rows():
    conn = get_conn()
    users = conn.execute("SELECT chat_id, language FROM users ORDER BY chat_id").fetchall()
    out = []
    for u in users:
        task_count = conn.execute("SELECT COUNT(*) c FROM tasks WHERE chat_id=?", (u["chat_id"],)).fetchone()["c"]
        times = conn.execute(
            "SELECT hour, minute FROM reminder_times WHERE chat_id=? ORDER BY hour, minute", (u["chat_id"],)
        ).fetchall()
        times_str = ", ".join(f"{r['hour']:02d}:{r['minute']:02d}" for r in times) or "—"
        out.append({"chat_id": u["chat_id"], "language": u["language"], "task_count": task_count, "times": times_str})
    conn.close()
    return out


def _daily_checkin_series(days=14):
    conn = get_conn()
    labels, done_counts, missed_counts = [], [], []
    for i in range(days - 1, -1, -1):
        d = (datetime.now(TZ) - timedelta(days=i)).strftime("%Y-%m-%d")
        done = conn.execute("SELECT COUNT(*) c FROM logs WHERE date=? AND done=1", (d,)).fetchone()["c"]
        missed = conn.execute("SELECT COUNT(*) c FROM logs WHERE date=? AND done=0", (d,)).fetchone()["c"]
        labels.append(d[5:])  # MM-DD
        done_counts.append(done)
        missed_counts.append(missed)
    conn.close()
    return {"labels": labels, "done": done_counts, "missed": missed_counts}


def _language_breakdown():
    conn = get_conn()
    rows = conn.execute("SELECT language, COUNT(*) c FROM users GROUP BY language").fetchall()
    conn.close()
    labels = {"km": "Khmer", "en": "English"}
    return {
        "labels": [labels.get(r["language"], r["language"]) for r in rows],
        "counts": [r["c"] for r in rows],
    }


@flask_app.route("/ping")
def ping():
    return Response("OK - Daily Habit Bot is running", mimetype="text/plain")


@flask_app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if secrets.compare_digest(request.form.get("password", ""), DASHBOARD_PASSWORD):
            session["logged_in"] = True
            session["csrf_token"] = secrets.token_hex(16)
            return redirect(url_for("dashboard"))
        error = "Wrong password."
    return render_template_string(LOGIN_HTML, css=DASHBOARD_CSS, error=error)


@flask_app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@flask_app.route("/")
@_require_login
def dashboard():
    return render_template_string(
        DASHBOARD_HTML,
        css=DASHBOARD_CSS,
        maintenance=is_maintenance_on(),
        db_synced=USE_TURSO,
        csrf_token=session.get("csrf_token", ""),
        stats=_dashboard_stats(),
        top_streaks=_top_streaks(),
        users=_user_rows(),
        checkin_series_json=json.dumps(_daily_checkin_series()),
        lang_breakdown_json=json.dumps(_language_breakdown()),
    )


@flask_app.route("/toggle-maintenance", methods=["POST"])
@_require_login
def toggle_maintenance():
    submitted = request.form.get("csrf_token", "")
    expected = session.get("csrf_token", "")
    if not expected or not secrets.compare_digest(submitted, expected):
        return Response("Invalid or expired form (CSRF check failed). Go back and try again.", status=400)
    set_setting("maintenance_mode", "false" if is_maintenance_on() else "true")
    return redirect(url_for("dashboard"))


def run_dashboard():
    port = int(os.environ.get("PORT", 10000))
    flask_app.run(host="0.0.0.0", port=port, threaded=True, use_reloader=False)


# ------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------
BOT_COMMANDS_EN = [
    BotCommand("start", "Register and see instructions"),
    BotCommand("menu", "Show the button menu"),
    BotCommand("addtask", "Add a daily task"),
    BotCommand("removetask", "Remove a task"),
    BotCommand("mytasks", "List your tasks and streaks"),
    BotCommand("addtime", "Add a daily reminder time"),
    BotCommand("removetime", "Remove a reminder time"),
    BotCommand("mytimes", "List your reminder times"),
    BotCommand("checkin", "Trigger a check-in now"),
    BotCommand("stats", "See your 7-day report"),
    BotCommand("export", "Download your data (Excel)"),
    BotCommand("language", "Switch language"),
    BotCommand("help", "Show help"),
]

BOT_COMMANDS_KM = [
    BotCommand("start", "ចុះឈ្មោះ និងមើលការណែនាំ"),
    BotCommand("menu", "បង្ហាញម៉ឺនុយប៊ូតុង"),
    BotCommand("addtask", "បន្ថែម task ប្រចាំថ្ងៃ"),
    BotCommand("removetask", "លុប task"),
    BotCommand("mytasks", "មើលបញ្ជី task និង streak"),
    BotCommand("addtime", "បន្ថែមម៉ោងជូនដំណឹង"),
    BotCommand("removetime", "លុបម៉ោងជូនដំណឹង"),
    BotCommand("mytimes", "មើលម៉ោងជូនដំណឹងទាំងអស់"),
    BotCommand("checkin", "សាកល្បង check-in ភ្លាមៗ"),
    BotCommand("stats", "របាយការណ៍ 7 ថ្ងៃចុងក្រោយ"),
    BotCommand("export", "ទាញយកទិន្នន័យ (Excel)"),
    BotCommand("language", "ប្តូរភាសា"),
    BotCommand("help", "មើលជំនួយ"),
]


async def post_init(application: Application):
    # Registers Telegram's native "☰" commands menu next to the text box.
    # Admin-only commands (e.g. /maintenance) are deliberately left out of
    # this public list.
    try:
        await application.bot.set_my_commands(BOT_COMMANDS_EN)  # fallback
        await application.bot.set_my_commands(BOT_COMMANDS_EN, language_code="en")
        await application.bot.set_my_commands(BOT_COMMANDS_KM, language_code="km")
    except Exception as exc:
        logger.warning("Failed to register the Telegram commands menu: %s", exc)
    await schedule_all_users(application)


async def error_handler(update, context: ContextTypes.DEFAULT_TYPE):
    """Logs unhandled exceptions instead of letting them vanish, and gives
    the user a friendly message instead of just going silent."""
    logger.error("Unhandled exception while processing update: %s", update, exc_info=context.error)
    try:
        chat_id = None
        if isinstance(update, Update):
            if update.effective_chat:
                chat_id = update.effective_chat.id
        if chat_id:
            lang = get_user_language(chat_id)
            msg = "⚠️ Something went wrong. Please try again." if lang != "km" else "⚠️ មានបញ្ហាកើតឡើង សូមសាកល្បងម្តងទៀត។"
            await context.bot.send_message(chat_id=chat_id, text=msg)
    except Exception:
        pass  # never let the error handler itself raise


def main():
    if BOT_TOKEN == "PUT_YOUR_TOKEN_HERE":
        env_keys = sorted(os.environ.keys())
        logger.error("BOT_TOKEN not found. Available env var keys: %s", env_keys)
        raise RuntimeError("Set the BOT_TOKEN environment variable before running.")

    init_db()

    if DASHBOARD_PASSWORD == "changeme":
        logger.warning("DASHBOARD_PASSWORD is not set — using the default 'changeme'. Set it in your environment!")

    threading.Thread(target=run_dashboard, daemon=True).start()

    application = Application.builder().token(BOT_TOKEN).post_init(post_init).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("menu", menu_command))
    application.add_handler(CommandHandler("addtask", add_task))
    application.add_handler(CommandHandler("removetask", remove_task))
    application.add_handler(CommandHandler("mytasks", my_tasks))
    application.add_handler(CommandHandler("addtime", add_time))
    application.add_handler(CommandHandler("removetime", remove_time))
    application.add_handler(CommandHandler("mytimes", my_times))
    application.add_handler(CommandHandler("checkin", manual_checkin))
    application.add_handler(CommandHandler("stats", stats))
    application.add_handler(CommandHandler("export", export_command))
    application.add_handler(CommandHandler("language", language_command))
    application.add_handler(CommandHandler("maintenance", maintenance_command))
    application.add_handler(CallbackQueryHandler(button_handler))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_input))
    application.add_error_handler(error_handler)

    logger.info("Bot starting...")

    try:
        asyncio.get_event_loop()
    except RuntimeError:
        asyncio.set_event_loop(asyncio.new_event_loop())

    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()